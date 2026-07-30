"""
竞品分析编排器
- 解析输入 Excel（「输入内容」sheet）
- 按数据源分组，通过 Crawshrimp HTTP API 提交任务
- 轮询任务状态，收集导出文件
- 合并结果
"""
import asyncio
import json
import logging
import os
import tempfile
import zipfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import httpx
import openpyxl

logger = logging.getLogger(__name__)

# 数据源配置
DATA_SOURCES = {
    "bi_dashboard": {
        "adapter_id": "semir-bi-dashboard",
        "task_id": "single_product_analysis",
        "label": "BI看板-单品分析",
        "input_key": "sku",  # 使用款号
        "output_prefix": "单品分析",
    },
    "product_360": {
        "adapter_id": "semir-product-360",
        "task_id": "product_data_export",
        "label": "商品360",
        "input_key": "sku",
        "output_prefix": "商品360",
    },
    "seller_qa": {
        "adapter_id": "tmall-seller-qa",
        "task_id": "ask_all_export",
        "label": "天猫问大家",
        "input_key": "self_id",  # 使用自品ID
        "output_prefix": "自品问大家",
    },
    "competitor": {
        "adapter_id": "tmall-ops-assistant",
        "task_id": "buyer_reviews",
        "label": "竞品评价&问大家",
        "input_key": "competitor_id",
        "output_prefix": "竞品评价",
    },
}


class InputRow:
    """输入内容中的一行数据"""
    def __init__(self, structure: str, sku: str, self_id: str, competitor_id: str, date_range: str):
        self.structure = structure
        self.sku = sku
        self.self_id = self_id
        self.competitor_id = competitor_id
        self.date_start, self.date_end = self._parse_date_range(date_range)

    @staticmethod
    def _parse_date_range(date_range: str) -> tuple:
        if not date_range or "~" not in date_range:
            return "", ""
        parts = date_range.split("~")
        return parts[0].strip(), parts[-1].strip()


class CompetitiveAnalysisOrchestrator:
    def __init__(self, api_base: str = "http://127.0.0.1:18765"):
        self.api_base = api_base.rstrip("/")
        self.runs: Dict[str, dict] = {}

    def parse_input_excel(self, file_path: str, sheet_name: str = "输入内容") -> List[InputRow]:
        """解析输入 Excel 文件"""
        wb = openpyxl.load_workbook(file_path, data_only=True)
        ws = wb[sheet_name]
        rows = []
        for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
            if not any(row):
                continue
            structure = str(row[0] or "").strip()
            sku = str(row[1] or "").strip() if row[1] else ""
            self_id = str(row[2] or "").strip() if row[2] else ""
            competitor_id = str(row[3] or "").strip() if row[3] else ""
            date_range = str(row[4] or "").strip() if row[4] else ""
            if not any([sku, self_id, competitor_id]):
                continue
            rows.append(InputRow(structure, sku, self_id, competitor_id, date_range))
        wb.close()
        return rows

    def _build_source_tasks(self, rows: List[InputRow]) -> Dict[str, dict]:
        """
        按数据源分组建构任务参数
        每个数据源一批处理（去重后）
        """
        tasks = {}

        # BI看板 & 商品360: 使用款号（去重，排除空）
        skus = list(set(r.sku for r in rows if r.sku))
        if skus:
            structure = rows[0].structure if rows else ""
            date_range = rows[0]
            tasks["bi_dashboard"] = {
                "adapter_id": "semir-bi-dashboard",
                "task_id": "single_product_analysis",
                "params": {
                    "sku_list": "\n".join(skus),
                    "date_start": date_range.date_start if date_range else "",
                    "date_end": date_range.date_end if date_range else "",
                    "structure_name": structure,
                },
            }
            tasks["product_360"] = {
                "adapter_id": "semir-product-360",
                "task_id": "product_data_export",
                "params": {
                    "sku_list": "\n".join(skus),
                    "structure_name": structure,
                },
            }

        # 天猫问大家: 使用自品ID
        self_ids = list(set(r.self_id for r in rows if r.self_id))
        if self_ids:
            structure = next((r.structure for r in rows if r.self_id), "")
            date_range = next((r for r in rows if r.self_id), None)
            tasks["seller_qa"] = {
                "adapter_id": "tmall-seller-qa",
                "task_id": "ask_all_export",
                "params": {
                    "item_ids": "\n".join(self_ids),
                    "date_start": date_range.date_start if date_range else "",
                    "date_end": date_range.date_end if date_range else "",
                    "structure_name": structure,
                },
            }

        # 竞品评价: 使用竞品ID构造商品链接
        competitor_ids = list(set(r.competitor_id for r in rows if r.competitor_id))
        if competitor_ids:
            links = [f"https://detail.tmall.com/item.htm?id={cid}" for cid in competitor_ids]
            tasks["competitor"] = {
                "adapter_id": "tmall-ops-assistant",
                "task_id": "buyer_reviews",
                "params": {
                    "item_links": "\n".join(links),
                },
            }

        return tasks

    async def _submit_task(self, http: httpx.AsyncClient, adapter_id: str, task_id: str, params: dict) -> Optional[dict]:
        """提交任务到 Crawshrimp API"""
        url = f"{self.api_base}/tasks/{adapter_id}/{task_id}/run"
        try:
            resp = await http.post(url, json=params, timeout=30)
            if resp.status_code == 200:
                return resp.json()
            logger.error(f"任务提交失败 {adapter_id}/{task_id}: {resp.status_code} {resp.text}")
            return None
        except Exception as e:
            logger.error(f"任务提交异常 {adapter_id}/{task_id}: {e}")
            return None

    async def _poll_task(self, http: httpx.AsyncClient, adapter_id: str, task_id: str, max_wait_s: int = 1800) -> dict:
        """轮询任务状态直到完成"""
        url = f"{self.api_base}/tasks/{adapter_id}/{task_id}/status"
        poll_interval = 10
        waited = 0
        while waited < max_wait_s:
            try:
                resp = await http.get(url, timeout=10)
                if resp.status_code == 200:
                    data = resp.json()
                    status = data.get("status", "running")
                    if status in ("done", "error", "stopped"):
                        return data
                await asyncio.sleep(poll_interval)
                waited += poll_interval
            except Exception as e:
                logger.warning(f"轮询异常 {adapter_id}/{task_id}: {e}")
                await asyncio.sleep(poll_interval)
                waited += poll_interval
        return {"status": "error", "error": "任务超时"}

    async def _get_export_data(self, http: httpx.AsyncClient, adapter_id: str, task_id: str) -> List[dict]:
        """获取任务的导出数据"""
        url = f"{self.api_base}/data/{adapter_id}/{task_id}/export"
        try:
            resp = await http.get(url, timeout=30)
            if resp.status_code == 200:
                return resp.json() if isinstance(resp.json(), list) else []
            return []
        except Exception as e:
            logger.error(f"获取导出数据失败 {adapter_id}/{task_id}: {e}")
            return []

    async def execute(self, file_path: str) -> str:
        """
        执行完整的竞品分析流程
        返回 run_id
        """
        run_id = f"ca_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.runs[run_id] = {
            "status": "running",
            "file_path": file_path,
            "progress": {},
            "results": {},
            "started_at": datetime.now().isoformat(),
        }

        try:
            # 1. 解析输入 Excel
            rows = self.parse_input_excel(file_path)
            if not rows:
                self.runs[run_id]["status"] = "error"
                self.runs[run_id]["error"] = "输入Excel无有效数据行"
                return run_id

            # 2. 建任务
            source_tasks = self._build_source_tasks(rows)
            total_sources = len(source_tasks)
            self.runs[run_id]["progress"] = {
                key: {"status": "pending", "completed": False}
                for key in source_tasks
            }

            async with httpx.AsyncClient() as http:
                # 3. 串行提交并等待（不同站点可并行，这里为简单串行）
                for source_key, task in source_tasks.items():
                    adapter_id = task["adapter_id"]
                    task_id = task["task_id"]
                    params = task["params"]
                    label = DATA_SOURCES.get(source_key, {}).get("label", source_key)

                    logger.info(f"[{run_id}] 开始取数: {label}")
                    self.runs[run_id]["progress"][source_key]["status"] = "running"

                    # 提交任务
                    submit_result = await self._submit_task(http, adapter_id, task_id, params)
                    if not submit_result:
                        self.runs[run_id]["progress"][source_key] = {
                            "status": "failed", "error": "任务提交失败"
                        }
                        continue

                    # 轮询等待
                    poll_result = await self._poll_task(http, adapter_id, task_id)
                    if poll_result.get("status") == "done":
                        self.runs[run_id]["progress"][source_key] = {
                            "status": "done", "completed": True
                        }
                        # 获取导出数据
                        export_data = await self._get_export_data(http, adapter_id, task_id)
                        self.runs[run_id]["results"][source_key] = export_data
                        logger.info(f"[{run_id}] 取数完成: {label}, {len(export_data)} 条")
                    else:
                        self.runs[run_id]["progress"][source_key] = {
                            "status": "failed",
                            "error": poll_result.get("error", "未知错误"),
                        }
                        logger.warning(f"[{run_id}] 取数失败: {label}")

                # 4. 汇总结果
                completed = sum(1 for p in self.runs[run_id]["progress"].values() if p.get("completed"))
                failed = sum(1 for p in self.runs[run_id]["progress"].values() if p.get("status") == "failed")
                self.runs[run_id]["status"] = "done"
                self.runs[run_id]["summary"] = {
                    "total_sources": total_sources,
                    "completed": completed,
                    "failed": failed,
                }

        except Exception as e:
            logger.exception(f"[{run_id}] 编排异常")
            self.runs[run_id]["status"] = "error"
            self.runs[run_id]["error"] = str(e)

        self.runs[run_id]["finished_at"] = datetime.now().isoformat()
        return run_id

    def get_status(self, run_id: str) -> Optional[dict]:
        run = self.runs.get(run_id)
        if not run:
            return None
        return {
            "run_id": run_id,
            "status": run["status"],
            "progress": run.get("progress", {}),
            "summary": run.get("summary"),
            "error": run.get("error"),
            "started_at": run.get("started_at"),
            "finished_at": run.get("finished_at"),
        }

    def export_merged(self, run_id: str, output_path: str) -> Optional[str]:
        """将各数据源结果合并导出为一个 Excel 文件"""
        run = self.runs.get(run_id)
        if not run or run["status"] != "done":
            return None

        wb = openpyxl.Workbook()
        wb.remove(wb.active)  # remove default sheet

        for source_key, rows in run.get("results", {}).items():
            if not rows:
                continue
            label = DATA_SOURCES.get(source_key, {}).get("label", source_key)
            sheet_name = label[:31]  # Excel sheet name max 31 chars
            ws = wb.create_sheet(title=sheet_name)

            # Write headers from first row's keys
            if rows:
                headers = list(rows[0].keys())
                ws.append(headers)
                for row in rows:
                    ws.append([row.get(h, "") for h in headers])

        # Add summary sheet
        ws_summary = wb.create_sheet(title="取数汇总")
        ws_summary.append(["数据源", "状态", "记录数"])
        for source_key, progress in run.get("progress", {}).items():
            label = DATA_SOURCES.get(source_key, {}).get("label", source_key)
            row_count = len(run.get("results", {}).get(source_key, []))
            ws_summary.append([label, progress.get("status", "未知"), row_count])

        os.makedirs(os.path.dirname(output_path) if os.path.dirname(output_path) else ".", exist_ok=True)
        wb.save(output_path)
        wb.close()
        return output_path


# 全局单例
_orchestrator: Optional[CompetitiveAnalysisOrchestrator] = None


def get_orchestrator(api_base: str = "http://127.0.0.1:18765") -> CompetitiveAnalysisOrchestrator:
    global _orchestrator
    if _orchestrator is None:
        _orchestrator = CompetitiveAnalysisOrchestrator(api_base=api_base)
    return _orchestrator
